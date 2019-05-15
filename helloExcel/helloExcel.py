# 创建xml
from openpyxl import Workbook

# 创建工作簿
wb = Workbook()

# 获取sheet页（默认1个）
ws = wb.active
# 取第二个sheet页(需要新建)
# sheet名字，位置
ws1 = wb.create_sheet(title='1', index=1)

# 设置sheet名字
ws.title = 'hello'

# 获取sheet的另一种方式
ws1 = wb['1']

# 获取sheet的方式
# ws1 = wb.get_sheet_by_name('1')

# print(wb.get_sheet_names())
print(wb.sheetnames)

ws['A1'] = 'hello A1'
ws['A2'] = 'hello A2'
ws['A3'] = 'hello A3'

print(ws.cell(row=1, column=1))
print(ws.cell(1, 2))
print(ws.cell(2, 2))
print(ws.cell(3, 4))
for row in ws.iter_rows(1, 3):
    for cell in row:
        cell.value = 'hello cell'

print(ws['A1':'C3'])

# print(ws.rows)
# print(ws.columns)
# 保存文件,如果存在会覆盖
wb.save('E:\hello.xlsx')

# 读取
from openpyxl import load_workbook

wb = load_workbook('E:\hello.xlsx')
ws = wb['hello']
for row in ws.rows:
    for cell in row:
        print(cell.value)
