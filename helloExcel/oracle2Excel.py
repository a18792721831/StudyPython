import cx_Oracle
import openpyxl

# 用户名
username = 'tj_20160217'
# 密码
password = 'tj_20160217'
# IP
ip = '10.0.250.19'
# 端口
port = '1521'
# 数据库实例名
servername = 'starbass'
# 获取连接
xlsxPath = 'E:\\sysconfigen.xlsx'
sheetName = 'sysconfigen'
wb = openpyxl.Workbook()
ws = wb.active
ws.title = sheetName
with cx_Oracle.connect(username + '/' + password + '@' + ip + ':' + port + '/' + servername) as db:
    cur = db.cursor()
    result = cur.execute('select * from sysconfigen')
    row = 1
    column = 1
    # 获取表头
    for header in cur.description:
        ws.cell(row, column).value = header[0]
        column += 1
    row += 1
    for one_result in result.__iter__():
        column = 1
        for at in one_result:
            if at:
                ws.cell(row, column).value = at
            else:
                # 空值处理
                ws.cell(row, column).value = 'null'
            column += 1
        row += 1
wb.save(xlsxPath)