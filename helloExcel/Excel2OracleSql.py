from openpyxl import load_workbook
from createSql import getSql


def excel2Sql():
    print('Please input file path:')
    filePath = input()
    excelFile = load_workbook(filePath)
    sheet = excelFile.active
    print('Please input author:')
    author = input()
    hebei = sheet.cell(1, 4).value
    liaoning = sheet.cell(1, 5).value
    heilongjiang = sheet.cell(1, 6).value
    yangquan = sheet.cell(1, 8).value
    file = [hebei, liaoning, heilongjiang, yangquan]
    for file_suffix in file:
        with open(filePath.replace('.xlsx', hebei + '.sql'), 'a') as file:
            for row in sheet.iter_rows(2, sheet.max_row):
                if file_suffix == hebei:
                    if row[3].fill.bgColor.index != 0:
                        file.write(getSql(int(row[0].value), str(row[1].value), str(row[2].value), author))
                if file_suffix == liaoning:
                    if row[4].fill.bgColor.index != 0:
                        file.write(getSql(int(row[0].value), str(row[1].value), str(row[2].value), author))
                if file_suffix == heilongjiang:
                    if row[5].fill.bgColor.index != 0:
                        file.write(getSql(int(row[0].value), str(row[1].value), str(row[2].value), author))
                if file_suffix == yangquan:
                    if row[6].fill.bgColor.index != 0:
                        file.write(getSql(int(row[0].value), str(row[1].value), str(row[2].value), author))

if __name__ == "__main__":
    excel2Sql()